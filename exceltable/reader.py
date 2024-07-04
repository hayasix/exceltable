#!/usr/bin/env python3
# vim: set fileencoding=utf-8 fileformat=unix expandtab :

import os
import datetime
from itertools import count
from collections import namedtuple, OrderedDict
import tempfile
import re

import openpyxl
from openpyxl.utils.cell import get_column_letter
import msoffcrypto


__all__ = ("Reader", "DictReader")


NEWLINE = "\n"
NOTIME = datetime.time(0)


class BaseReader(object):

    """Reader for a table on a Excel sheet"""

    def __init__(self, source: str,
            sheet: str = None, password: str = None,
            start_row=0, stop_row=None, start_col=0, stop_col=None,
            header_rows=1, empty=None, repeat=False, trim=True):
        """Initiator.

        Parameters
        ----------
        source : openpyxl.workbook.workbook.Workbook, str, PathLike
            Excel book, or pathname of Excel book (.xl*)
        sheet : str or None
            worksheet name; None=leftmost
        password : str
            password for encrypted Excel book
        start_row : int
            top row number; starts with 0
        stop_row : int, float, str, callable
            (int) bottom row number [1]_ (base=0)
            (float, str) boundary marker string to stop scan
            (callable) see below [2]_
        start_col : int
            left column number; starts with 0
        stop_col : int, float, str, callable
            (int) right column number [3]_ ; starts with 0
            (float or str) boundary marker string to stop scan
            (callable) see below [4]_
        header_rows : int
            rows to read as field names; default=1
        empty : float, str
            alternative value for empty cells
        repeat : bool
            repeat cell value of the previous row if blank
        trim : bool
            suppress redundant zeros [5]_; default=True

        .. [1]  stop_row itself is not included in the data read.
        .. [2]  this callable (function) is to accept a list with the row
                values and to return True if the row is the boundary.
        .. [3]  stop_col itself is not included in the data read.
        .. [4]  this callable (function) is to accept a (cell) value and
                to return True if the column of the cell is the boundary.
        .. [5]  e.g. 3.0 -> 3, 2000-12-31 00:00:00 -> 2000-12-31
        """
        self.tempfile = None
        if isinstance(source, (str, os.PathLike)):
            source = str(source)
            if password:
                self.tempfile_id, self.tempfile = tempfile.mkstemp(suffix=".xlsx")
                with (open(source, "rb") as in_,
                      open(self.tempfile, "wb") as out):
                    f = msoffcrypto.OfficeFile(in_)
                    f.load_key(password=password)
                    f.decrypt(out)
                source = self.tempfile
            self.book = openpyxl.load_workbook(source, data_only=True)
        else:
            self.book = source
        if sheet and isinstance(sheet, str):
            self.sheet = self.book[sheet]
        else:
            self.sheet = self.book.worksheets[sheet or 0]
        self.start_row = start_row
        self.stop_row = stop_row
        self.start_col = start_col
        self.stop_col = stop_col
        self.header_rows = header_rows
        self.empty = empty
        self.repeat = repeat
        self.trim = trim
        self._rows = self.sheet.iter_rows(min_row=start_row + 1,
                                          min_col=start_col + 1,
                                          values_only=True)
        self.fieldnames = self._get_fields(header_rows)

    def _mergearea(self, row, col):
        """Get the associated merge area.

        Parameters
        ----------
        row : int
            row number of a cell (base=0)
        col : int
            column number of a cell (base=0)

        Returns
        -------
        tuple[int]
            a tuple of 4 integers as (rlo, rhi, clo, chi), where:
                rlo:  min. row number (base=0)
                rhi:  max. row number + 1 (base=0)
                clo:  min. column number (base=0)
                chi:  max. column number + 1 (base=0)
        """
        row, col = row + 1, col + 1
        for r in self.sheet.merged_cells.ranges:
            if (r.min_row <= row <= r.max_row and
                r.min_col <= col <= r.max_col):
                return (r.min_row - 1, r.max_row - 1,
                        r.min_col - 1, r.max_col - 1)
        return None

    @staticmethod
    def _isbreak_factory(criteria):
        if isinstance(criteria, int):
            def isbreak(row_or_col, _): return row_or_col == criteria
        elif isinstance(criteria, (float, str)):
            def isbreak(_, value): return value == criteria
        elif callable(criteria):
            def isbreak(_, value): return criteria(value)
        else:
            def isbreak(_, value): return not(bool(value))
        return isbreak

    def _get_fields(self, header_rows=1):
        """Get the field names.

        Parameters
        ----------
        header_rows : int
            number of rows of the header [default: 1]

        Returns
        -------
        list[str]
            field names

        Notes
        -----
        If 1 < header_rows, each field name consists of the values of
        vertically continuous cells joined by '_' (underscore).
        """
        fields = []
        rows = [next(self._rows) for _ in range(header_rows)]
        isbreak = self._isbreak_factory(self.stop_col)
        for col in range(max(len(r) for r in rows)):
            abscol = self.start_col + col
            f = []
            for row in range(header_rows):
                absrow = self.start_row + row
                ma = self._mergearea(absrow, abscol)
                if not ma:
                    try:
                        v = rows[row][col]
                    except IndexError:
                        v = ""
                elif ma[0] == absrow:
                    v = rows[row][ma[2]]
                else:
                    v = None
                if v:
                    v = str(v)
                    if v.endswith(".0"):
                        v = v[:-2]
                    f.append(v)
            field = "_".join(f).replace(NEWLINE, "")
            fields.append(field or get_column_letter(abscol + 1))
            if isbreak(abscol, field):
                fields.pop()
                break
        # Add subscriptions for duplicate field names.
        for k, v in enumerate(fields):
            if v not in fields[:k]: continue
            for n in count(1):
                alt = f"{v}_{n}"
                if alt not in fields[:k]:
                    fields[k] = alt
                    break
        return fields

    @staticmethod
    def _trim(values):
        for v in values:
            if isinstance(v, float) and v.is_integer():
                v = int(v)
            elif isinstance(v, datetime.datetime) and v.time() == NOTIME:
                v = v.date()
            yield v

    def _build(self, kv):
        raise NotImplementedError

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        pass

    def __del__(self):
        if self.tempfile:
            os.close(self.tempfile_id)
            os.remove(self.tempfile)
        else:
            self.book.close()
        # To avoid openpyxl bug:
        import gc; gc.collect()

    def __iter__(self):
        isbreak = self._isbreak_factory(self.stop_row)
        if self.repeat: prev = [None] * len(self.fieldnames)
        for absrow, row in enumerate(self._rows,
                             start=self.start_row + self.header_rows):
            values = [self.empty if v is None else v for v in row]
            if len(values) < 1 or isbreak(absrow, values[0]): break
            if self.repeat:
                values = [p if v in (None, "") else v
                          for (v, p) in zip(values, prev)]
            if self.trim: values = self._trim(values)
            yield self._build(self.fieldnames, values)
            if self.repeat: prev = values


class Reader(BaseReader):

    def _build(self, keys, values):
        try:
            return self.CSVRecord(*values)
        except AttributeError:
            keys = [re.sub(r"\W", "_", k) for k in keys]
            self.CSVRecord = namedtuple("CSVRecord", keys)
            return self.CSVRecord(*values)


class DictReader(BaseReader):

    def _build(self, keys, values):
        return OrderedDict(zip(keys, values))
